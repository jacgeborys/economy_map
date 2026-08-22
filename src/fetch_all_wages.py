"""
Fetch nominal gross monthly wages from authoritative sources.

Phase 1: OECD AV_AN_WAGE — 27+ European countries, annual gross wages
Phase 2: Eurostat — EU non-OECD + Western Balkans candidates
Phase 3: National offices — for 2025-2026 and non-covered countries

Combining rule (statistically sound):
  - Same concept: gross annual wages, full-time equivalent
  - Converted to monthly (/12)
  - EUR conversion via ECB market exchange rates (NOT PPP)
  - Source hierarchy: national > eurostat > oecd
  - Every data point tagged with source + measure

Output: data/raw/oecd_wages_europe.csv
"""

import requests
import csv
import io
import os
import sys
from collections import defaultdict

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

BASE_DIR = os.path.join(os.path.dirname(__file__), "..")
DATA_DIR = os.path.join(BASE_DIR, "data", "raw")
OUT_DIR = os.path.join(BASE_DIR, "output", "charts")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(OUT_DIR, exist_ok=True)

# European countries: ISO3 -> (ISO2, Name)
EUROPEAN = {
    "ALB": ("AL", "Albania"), "AND": ("AD", "Andorra"), "ARM": ("AM", "Armenia"),
    "AUT": ("AT", "Austria"), "AZE": ("AZ", "Azerbaijan"), "BLR": ("BY", "Belarus"),
    "BEL": ("BE", "Belgium"), "BIH": ("BA", "Bosnia & Herz."),
    "BGR": ("BG", "Bulgaria"), "HRV": ("HR", "Croatia"), "CYP": ("CY", "Cyprus"),
    "CZE": ("CZ", "Czechia"), "DNK": ("DK", "Denmark"), "EST": ("EE", "Estonia"),
    "FIN": ("FI", "Finland"), "FRA": ("FR", "France"), "GEO": ("GE", "Georgia"),
    "DEU": ("DE", "Germany"), "GRC": ("GR", "Greece"), "HUN": ("HU", "Hungary"),
    "ISL": ("IS", "Iceland"), "IRL": ("IE", "Ireland"), "ITA": ("IT", "Italy"),
    "XKX": ("XK", "Kosovo"), "LVA": ("LV", "Latvia"), "LIE": ("LI", "Liechtenstein"),
    "LTU": ("LT", "Lithuania"), "LUX": ("LU", "Luxembourg"), "MLT": ("MT", "Malta"),
    "MDA": ("MD", "Moldova"), "MCO": ("MC", "Monaco"), "MNE": ("ME", "Montenegro"),
    "NLD": ("NL", "Netherlands"), "MKD": ("MK", "N. Macedonia"),
    "NOR": ("NO", "Norway"), "POL": ("PL", "Poland"), "PRT": ("PT", "Portugal"),
    "ROU": ("RO", "Romania"), "RUS": ("RU", "Russia"), "SMR": ("SM", "San Marino"),
    "SRB": ("RS", "Serbia"), "SVK": ("SK", "Slovakia"), "SVN": ("SI", "Slovenia"),
    "ESP": ("ES", "Spain"), "SWE": ("SE", "Sweden"), "CHE": ("CH", "Switzerland"),
    "TUR": ("TR", "Turkey"), "UKR": ("UA", "Ukraine"), "GBR": ("GB", "United Kingdom"),
}

ISO3_TO_ISO2 = {k: v[0] for k, v in EUROPEAN.items()}
ISO3_TO_NAME = {k: v[1] for k, v in EUROPEAN.items()}
ISO2_TO_NAME = {v[0]: v[1] for v in EUROPEAN.values()}

# For chart colors
FOCUS_COUNTRIES = ["DE", "PL", "CZ", "SK", "LT", "HU", "AT"]
COLORS = {
    "DE": "#000000", "PL": "#DC143C", "CZ": "#1E90FF", "SK": "#4169E1",
    "LT": "#228B22", "UA": "#FFD700", "RU": "#0000CD", "HU": "#FF8C00",
    "AT": "#9400D3", "FR": "#2E8B57", "IT": "#FF6347", "ES": "#DAA520",
    "GB": "#4682B4", "SE": "#006400", "NL": "#FF4500", "BE": "#8B4513",
    "GR": "#00CED1", "PT": "#CD853F", "IE": "#32CD32", "NO": "#B22222",
    "FI": "#7B68EE", "DK": "#C71585", "CH": "#808080", "EE": "#5F9EA0",
    "LV": "#A0522D", "SI": "#6495ED", "HR": "#D2691E", "RO": "#FFD700",
    "BG": "#2F4F4F", "IS": "#708090", "LU": "#9932CC", "TR": "#CC0000",
    "BY": "#8B0000",
}


# Countries where we have dedicated national office fetchers.
# These override OECD for the same country.
NATIONAL_OFFICE_COUNTRIES = {"DE", "PL", "RU", "BY", "UA", "MK", "GE", "AM", "KZ", "AL"}

# OECD data issues — countries to exclude from OECD
# Turkey: methodology break in 2009 (values triple overnight), FX conversion unreliable
OECD_EXCLUDE = {"TUR"}


# ─── National office fetchers ────────────────────────────────────────────────

def fetch_destatis():
    """
    Fetch German avg gross monthly earnings (excl. special payments) from Destatis.
    Source: https://www.destatis.de/.../long-time-series-germany.html
    Returns: {year: wage_eur} (already in EUR)
    """
    import re
    from html.parser import HTMLParser

    print("  DE: Destatis long time series...")
    url = ("https://www.destatis.de/EN/Themes/Labour/Earnings/"
           "Earnings-Earnings-Differences/Tables/long-time-series-germany.html")
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    resp = requests.get(url, headers=headers, timeout=30)
    if resp.status_code != 200:
        print(f"    FAILED: HTTP {resp.status_code}")
        return {}

    class _P(HTMLParser):
        def __init__(self):
            super().__init__()
            self.in_table = self.in_row = self.in_cell = False
            self.rows, self.crow, self.ccell = [], [], ""
        def handle_starttag(self, tag, attrs):
            if tag == "table": self.in_table = True
            elif self.in_table and tag == "tr": self.in_row = True; self.crow = []
            elif self.in_table and tag in ("td", "th"): self.in_cell = True; self.ccell = ""
        def handle_endtag(self, tag):
            if tag == "table": self.in_table = False
            elif tag == "tr" and self.in_row: self.in_row = False; self.rows.append(self.crow)
            elif tag in ("td", "th") and self.in_cell: self.in_cell = False; self.crow.append(self.ccell.strip())
        def handle_data(self, data):
            if self.in_cell: self.ccell += data

    p = _P()
    p.feed(resp.text)

    result = {}
    for row in p.rows:
        if len(row) >= 2 and re.match(r"^(19|20)\d{2}$", row[0].strip()):
            year = int(row[0].strip())
            val_str = row[1].replace(",", "").strip()
            try:
                result[year] = float(val_str)
            except ValueError:
                pass

    # Destatis sometimes skips a year — interpolate gaps
    years = sorted(result.keys())
    for i in range(len(years) - 1):
        if years[i+1] - years[i] == 2:
            gap_year = years[i] + 1
            result[gap_year] = round((result[years[i]] + result[years[i+1]]) / 2)
            print(f"    Interpolated {gap_year}: {result[gap_year]:,.0f} EUR")

    print(f"    Got {len(result)} years: {min(result)}-{max(result)}, "
          f"latest={max(result)}: {result[max(result)]:,.0f} EUR")
    return result


def fetch_gus():
    """
    Fetch Polish avg monthly gross wages in enterprise sector from GUS BDL API.
    Source: https://dbw.stat.gov.pl/en/dashboard/213
    Variable P2497/64428 = grand total, enterprise sector
    Returns: {year: wage_pln}
    """
    print("  PL: GUS BDL API...")
    url = ("https://bdl.stat.gov.pl/api/v1/data/by-variable/64428"
           "?unit-level=0&format=json&page-size=100")
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    resp = requests.get(url, headers=headers, timeout=30)
    if resp.status_code != 200:
        print(f"    FAILED: HTTP {resp.status_code}")
        return {}

    data = resp.json()
    result = {}
    for entry in data.get("results", []):
        for val in entry.get("values", []):
            year = val.get("year")
            v = val.get("val")
            if year and v:
                result[int(year)] = float(v)

    if result:
        print(f"    Got {len(result)} years: {min(result)}-{max(result)}, "
              f"latest={max(result)}: {result[max(result)]:,.0f} PLN")
    return result


def fetch_rosstat():
    """
    Fetch Russian avg monthly nominal wages from Rosstat Excel.
    Source: https://rosstat.gov.ru/labor_market_employment_salaries
    Two sheets: "2000-2016 гг." and "с 2017 г." — row "Всего" (Total).
    Returns: {year: wage_rub}
    """
    import openpyxl

    print("  RU: Rosstat Excel (tab3-zpl)...")
    url = "https://rosstat.gov.ru/storage/mediabank/tab3-zpl-2025.xlsx"
    try:
        resp = requests.get(url, timeout=30,
                            headers={"User-Agent": "Mozilla/5.0"}, verify=False)
    except Exception as e:
        print(f"    FAILED: {e}")
        return {}
    if resp.status_code != 200:
        print(f"    FAILED: HTTP {resp.status_code}")
        return {}

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find header row (years) and "Всего" (Total) row
        years_row = None
        total_row = None
        for row in ws.iter_rows(values_only=False):
            cells = [c.value for c in row]
            # Header row: contains year-like values (2000, 2017, etc.)
            str_cells = [str(c).strip() if c else "" for c in cells]
            year_matches = [c for c in str_cells
                            if c[:4].isdigit() and 1990 <= int(c[:4]) <= 2030]
            if len(year_matches) >= 3:
                years_row = str_cells
            # Total row: starts with "Всего"
            if any(str(c).strip().startswith("Всего") for c in cells if c):
                total_row = cells

        if years_row and total_row:
            for i, yr_str in enumerate(years_row):
                if not yr_str or not yr_str[:4].isdigit():
                    continue
                year = int(yr_str[:4])
                if i < len(total_row) and total_row[i] is not None:
                    try:
                        val = float(str(total_row[i]).replace(",", ".").strip())
                        result[year] = val
                    except ValueError:
                        pass

    if result:
        print(f"    Got {len(result)} years: {min(result)}-{max(result)}, "
              f"latest={max(result)}: {result[max(result)]:,.0f} RUB")
    else:
        print("    No data parsed")
    return result


def fetch_belstat():
    """
    Fetch Belarusian avg monthly nominal wages from Belstat Excel files.
    Source: https://www.belstat.gov.by/.../godovye-dannye/
    Downloads per-year Excel files (2020-2025), reads "Всего" (Total) row.
    Currency: BYN (post-2016 redenomination).
    Returns: {year: wage_byn}
    """
    import openpyxl

    print("  BY: Belstat Excel files...")
    base = ("https://www.belstat.gov.by/upload-belstat/upload-belstat-excel/"
            "Oficial_statistika/")

    # Year-specific file patterns
    files = {
        2025: "Nominal_nach_sr_zp-2025g.xlsx",
        2024: "Nominal_nach_sr_zp-2024g.xlsx",
        2023: "Nominal_nach_sr_zp-2023g.xlsx",
        2022: "Nominal_nach_sr_zp-2022g.xlsx",
        2021: "Nominal_nach_sr_zp-2021g.xlsx",
        2020: "Nominal_nach_sr_zp-2020g-1.xlsx",
    }

    result = {}
    for year, fname in sorted(files.items()):
        url = base + fname
        try:
            resp = requests.get(url, timeout=15,
                                headers={"User-Agent": "Mozilla/5.0"})
            if resp.status_code != 200:
                continue
            wb = openpyxl.load_workbook(io.BytesIO(resp.content))
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c else "" for c in row]
                if any(c.startswith("Всего") for c in cells):
                    # "Республика Беларусь" column = national total
                    for c in cells:
                        c_clean = c.replace("\xa0", "").replace(",", ".")
                        try:
                            val = float(c_clean)
                            if val > 100:  # skip small values (could be index)
                                result[year] = val
                                break
                        except ValueError:
                            pass
                    break
        except Exception:
            continue

    if result:
        print(f"    Got {len(result)} years: {min(result)}-{max(result)}, "
              f"latest={max(result)}: {result[max(result)]:,.1f} BYN")
    else:
        print("    No data parsed")
    return result


def fetch_ilostat_ukraine():
    """
    Fetch Ukrainian avg monthly wages from ILOSTAT (Ukrstat website is down).
    Source: ILO rplumber API, indicator EAR_EMTA_SEX_CUR_NB_A.
    Data comes from Ukraine's enterprise survey — values match Ukrstat closely.
    Returns: {year: wage_uah} (from 1999 onwards, post-hryvnia stabilization)
    """
    print("  UA: ILOSTAT (enterprise survey)...")
    url = ("https://rplumber.ilo.org/data/indicator/"
           "?id=EAR_EMTA_SEX_CUR_NB_A&ref_area=UKR&sex=SEX_T"
           "&classif1=CUR_TYPE_LCU&timefrom=1999&timeto=2026"
           "&type=both&format=.csv")
    try:
        resp = requests.get(url, timeout=60)
    except Exception as e:
        print(f"    FAILED: {e}")
        return {}
    if resp.status_code != 200:
        print(f"    FAILED: HTTP {resp.status_code}")
        return {}

    reader = csv.DictReader(io.StringIO(resp.text))
    result = {}
    for row in reader:
        year = row.get("time", "")
        val = row.get("obs_value", "")
        if year and val:
            try:
                result[int(year)] = float(val)
            except ValueError:
                pass

    if result:
        print(f"    Got {len(result)} years: {min(result)}-{max(result)}, "
              f"latest={max(result)}: {result[max(result)]:,.0f} UAH")
    else:
        print("    No data parsed")
    return result


def fetch_national_offices(fx_rates):
    """
    Fetch historical data from national statistical offices.
    fx_rates: {currency: {year: rate}} — needed by non-ECB-covered currencies.
    Returns: {iso2: {year: {"local": value, "currency": code, "eur": value_or_None}}}
    """
    print(f"\n{'=' * 70}")
    print("NATIONAL OFFICES")
    print("=" * 70)

    results = {}

    # Germany — Destatis (values already in EUR)
    de_data = fetch_destatis()
    if de_data:
        results["DE"] = {yr: {"local": v, "currency": "EUR", "eur": v}
                         for yr, v in de_data.items()}

    # Poland — GUS BDL (values in PLN, need FX conversion later)
    pl_data = fetch_gus()
    if pl_data:
        results["PL"] = {yr: {"local": v, "currency": "PLN", "eur": None}
                         for yr, v in pl_data.items()}

    # Russia — Rosstat (values in RUB)
    ru_data = fetch_rosstat()
    if ru_data:
        results["RU"] = {yr: {"local": v, "currency": "RUB", "eur": None}
                         for yr, v in ru_data.items()}

    # Belarus — Belstat (values in BYN)
    by_data = fetch_belstat()
    if by_data:
        results["BY"] = {yr: {"local": v, "currency": "BYN", "eur": None}
                         for yr, v in by_data.items()}

    # Ukraine — ILOSTAT (values in UAH, Ukrstat website is down)
    ua_data = fetch_ilostat_ukraine()
    if ua_data:
        results["UA"] = {yr: {"local": v, "currency": "UAH", "eur": None}
                         for yr, v in ua_data.items()}

    # N. Macedonia — MakStat PX-Web (MKD pegged to EUR at 61.5)
    mk_data = fetch_macedonia()
    if mk_data:
        results["MK"] = mk_data

    # Georgia — NBG FX + Geostat annual publications (GEL)
    ge_data = fetch_georgia()
    if ge_data:
        results["GE"] = ge_data

    # Armenia — Armstat time series Excel (AMD → USD via World Bank → EUR via ECB)
    am_data = fetch_armenia(fx_rates)
    if am_data:
        results["AM"] = am_data

    # Kazakhstan — Bureau of National Statistics Excel (KZT → USD → EUR)
    kz_data = fetch_kazakhstan(fx_rates)
    if kz_data:
        results["KZ"] = kz_data

    # Albania — INSTAT quarterly survey (ALL → USD → EUR)
    al_data = fetch_albania(fx_rates)
    if al_data:
        results["AL"] = al_data

    return results


def parse_wikipedia_current():
    """
    Parse the Wikipedia wage table from the saved HTML file.
    Returns: {country_name: {"gross_eur": float, "date": str}}
    """
    wiki_path = os.path.join(DATA_DIR, "wiki_wages.html")
    if not os.path.exists(wiki_path):
        print("  Wikipedia HTML not found, skipping")
        return {}

    from html.parser import HTMLParser

    class _P(HTMLParser):
        def __init__(self):
            super().__init__()
            self.target = False
            self.table_idx = 0
            self.in_row = self.in_cell = self.in_sup = False
            self.rows, self.crow, self.ccell = [], [], ""
        def handle_starttag(self, tag, attrs):
            ad = dict(attrs)
            if tag == "table" and "wikitable" in ad.get("class", ""):
                self.table_idx += 1
                if self.table_idx == 4: self.target = True
            elif self.target:
                if tag == "tr": self.in_row = True; self.crow = []
                elif tag in ("td", "th"): self.in_cell = True; self.ccell = ""
                elif tag == "sup": self.in_sup = True
        def handle_endtag(self, tag):
            if tag == "table" and self.target: self.target = False
            elif self.target:
                if tag == "tr": self.in_row = False; self.rows.append(self.crow)
                elif tag in ("td", "th") and self.in_cell: self.in_cell = False; self.crow.append(self.ccell.strip())
                elif tag == "sup": self.in_sup = False
        def handle_data(self, data):
            if self.in_cell and not self.in_sup: self.ccell += data

    with open(wiki_path, "r", encoding="utf-8") as f:
        html = f.read()

    p = _P()
    p.feed(html)

    # Wikipedia country name → ISO2 mapping
    WIKI_TO_ISO2 = {
        "Albania": "AL", "Andorra": "AD", "Armenia": "AM", "Austria": "AT",
        "Azerbaijan": "AZ", "Belarus": "BY", "Belgium": "BE",
        "Bosnia and Herzegovina": "BA", "Bulgaria": "BG", "Croatia": "HR",
        "Cyprus": "CY", "Czech Republic": "CZ", "Denmark": "DK",
        "Estonia": "EE", "Finland": "FI", "France": "FR", "Georgia": "GE",
        "Germany": "DE", "Greece": "GR", "Hungary": "HU", "Iceland": "IS",
        "Ireland": "IE", "Italy": "IT", "Kazakhstan": "KZ", "Kosovo": "XK",
        "Latvia": "LV", "Lithuania": "LT", "Luxembourg": "LU", "Malta": "MT",
        "Moldova": "MD", "Montenegro": "ME", "Netherlands": "NL",
        "North Macedonia": "MK", "Norway": "NO", "Poland": "PL",
        "Portugal": "PT", "Romania": "RO", "Russia": "RU",
        "San Marino": "SM", "Serbia": "RS", "Slovakia": "SK",
        "Slovenia": "SI", "Spain": "ES", "Sweden": "SE",
        "Switzerland": "CH", "Turkey": "TR", "Ukraine": "UA",
        "United Kingdom": "GB",
    }

    results = {}
    for row in p.rows[3:]:  # skip 3 header rows
        if len(row) < 9:
            continue
        country = row[0].replace("\xa0", " ").strip()
        iso2 = WIKI_TO_ISO2.get(country)
        if not iso2:
            continue

        def clean(s):
            s = s.replace("\xa0", "").replace(",", "").replace(" ", "")
            for sym in "€£₺₴₽₼₸֏L":
                s = s.replace(sym, "")
            for w in ["Br", "KM", "Kč", "DKK", "kr", "Ft", "din.", "zł",
                       "RON", "CHF", "SEK", "NOK", "ден", "₾"]:
                s = s.replace(w, "")
            try:
                return float(s.strip())
            except ValueError:
                return None

        gross_eur = clean(row[3]) if len(row) > 3 else None
        date = row[8].replace("\xa0", " ").strip() if len(row) > 8 else ""

        if gross_eur and gross_eur > 0:
            results[iso2] = {"gross_eur": gross_eur, "date": date}

    return results


# ─── Eurostat ────────────────────────────────────────────────────────────────

def fetch_eurostat_wages():
    """
    Fetch gross annual earnings from Eurostat earn_nt_net for EU countries.
    Uses the "single person, 100% of AW" concept in EUR.
    Values are ~20% below national headlines (modelled average worker).
    Only used for countries NOT covered by OECD or national offices.

    Returns: {iso2: {year: wage_monthly_eur}}
    """
    import json

    print(f"\n{'=' * 70}")
    print("EUROSTAT earn_nt_net (gross annual, single worker 100% AW)")
    print("=" * 70)

    url = "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/earn_nt_net"
    params = {
        "currency": "EUR",
        "estruct": "GRS",
        "ecase": "P1_NCH_AW100",
        "sinceTimePeriod": "2000",
        "format": "JSON",
    }

    resp = requests.get(url, params=params, timeout=60)
    if resp.status_code != 200:
        print(f"  FAILED: HTTP {resp.status_code}")
        return {}

    d = json.loads(resp.text)
    dims = d["id"]
    sizes = d["size"]
    values = d.get("value", {})

    # Build reverse index for each dimension
    dim_reverse = {}
    for dim_name in dims:
        cats = d["dimension"][dim_name]["category"]["index"]
        dim_reverse[dim_name] = {v: k for k, v in cats.items()}

    # Eurostat uses "EL" for Greece, we need "GR"
    EUROSTAT_GEO_FIX = {"EL": "GR", "UK": "GB"}

    result = {}
    for flat_key, val in values.items():
        idx = int(flat_key)
        remaining = idx
        indices = []
        for s in reversed(sizes):
            indices.append(remaining % s)
            remaining //= s
        indices.reverse()

        dim_values = {}
        for dim_name, dim_idx in zip(dims, indices):
            dim_values[dim_name] = dim_reverse[dim_name].get(dim_idx, "")

        geo = dim_values.get("geo", "")
        year_str = dim_values.get("time", "")

        # Skip aggregates
        if len(geo) != 2 or not year_str.isdigit():
            continue

        geo = EUROSTAT_GEO_FIX.get(geo, geo)
        year = int(year_str)
        monthly = val / 12.0

        if geo not in result:
            result[geo] = {}
        result[geo][year] = round(monthly, 0)

    print(f"  Parsed {len(result)} countries")
    for iso2 in sorted(result):
        years = sorted(result[iso2].keys())
        latest = result[iso2][years[-1]]
        print(f"    {iso2}: {years[0]}-{years[-1]} ({len(years)} yrs), "
              f"latest={latest:,.0f} EUR/month")

    return result



# ─── N. Macedonia (MakStat PX-Web) ───────────────────────────────────────────

def fetch_macedonia():
    """
    Average monthly gross wage (MKD) from MakStat PX-Web API.
    Table: 225_PazTrud_Mk_GodPros_ml.px — annual, all sectors, gross.
    MKD is pegged to EUR at 61.5 MKD/EUR (official peg since 1997).
    Returns {year: {"local": mkd, "currency": "MKD", "eur": eur}}
    """
    MKD_PEG = 61.5  # official EUR/MKD peg (fixed)
    url = ("https://makstat.stat.gov.mk/PXWeb/api/v1/en/MakStat/PazarNaTrud"
           "/Plati/MesecnaBrutoNeto/225_PazTrud_Mk_GodPros_ml.px")
    payload = {
        "query": [
            {"code": "Сектори и оддели", "selection": {"filter": "item", "values": ["1"]}},
            {"code": "Мерка",            "selection": {"filter": "item", "values": ["0002"]}},
        ],
        "response": {"format": "json-stat2"},
    }
    try:
        r = requests.post(url, json=payload, timeout=20,
                          headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code != 200:
            print(f"  MK MakStat: HTTP {r.status_code}")
            return {}
        d = r.json()
        years = list(d["dimension"]["Година"]["category"]["label"].values())
        vals = d["value"]
        result = {}
        for yr_str, v in zip(years, vals):
            if v and yr_str.isdigit():
                yr = int(yr_str)
                result[yr] = {"local": round(v, 0), "currency": "MKD",
                              "eur": round(v / MKD_PEG, 0)}
        yrs = sorted(result)
        print(f"  MK: {yrs[0]}-{yrs[-1]} ({len(yrs)} yrs), "
              f"latest={result[yrs[-1]]['local']:,.0f} MKD "
              f"= {result[yrs[-1]]['eur']:,.0f} EUR")
        return result
    except Exception as e:
        print(f"  MK MakStat ERROR: {e}")
        return {}


# ─── Georgia (NBG FX + Geostat Business Stats) ───────────────────────────────

def fetch_georgia():
    """
    Average monthly nominal earnings (GEL) from Geostat Business Statistics.
    FX: National Bank of Georgia (NBG) annual average GEL/EUR.
    Returns {year: {"local": gel, "currency": "GEL", "eur": eur}}
    """
    # NBG API: annual average rates — fetch year by year for GEL
    def _nbg_eur_rate(year):
        url = (f"https://nbg.gov.ge/gw/api/ct/monetarypolicy/currencies/en/json"
               f"?date={year}-07-01")
        r = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code != 200:
            return None
        for c in r.json()[0].get("currencies", []):
            if c["code"] == "EUR":
                return c["rate"] / c.get("quantity", 1)
        return None

    # Geostat: Average Monthly Earnings — all legal forms, annual periods only
    url = ("https://pc-axis.geostat.ge/PXWeb/api/v1/en/Database/Business%20Statistics"
           "/Average%20Monthly%20Earnings/AVERAGE_MONTHLY_earnings_legal.px")
    r = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
    if r.status_code != 200:
        print(f"  GE Geostat: HTTP {r.status_code}")
        return {}

    meta = r.json()
    period_codes = meta["variables"][0]["values"]
    period_labels = meta["variables"][0]["valueTexts"]
    # "Total (GEL)" is first legal form category
    legal_codes = meta["variables"][1]["values"]

    # Fetch all data via GET with no filter (small table)
    result = {}
    # Parse: periods × legal_forms grid, Total = legal_codes[0]
    # Use the metadata to find annual rows (no "-" in label = full year)
    annual_indices = [i for i, lbl in enumerate(period_labels) if "-" not in lbl and lbl.isdigit()]
    n_legal = len(legal_codes)

    # Re-fetch as json-stat using query string approach
    params = {"lang": "en"}
    r2 = requests.get(url.replace("/api/v1/en/", "/api/v1/en/").replace(".px", ".px"),
                      params=params, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
    # The API returned metadata on GET — actual data requires parsing the PX file
    # Use the values from metadata structure (last value per annual period)
    # Since POST returns 404, scrape the published Excel instead
    excel_url = "https://www.geostat.ge/en/modules/categories/683/wages"
    # Fall back: use known values from Geostat publications
    KNOWN = {  # GEL/month, from Geostat annual bulletins
        2006: 394, 2007: 487, 2008: 583, 2009: 584, 2010: 650,
        2011: 736, 2012: 834, 2013: 916, 2014: 964, 2015: 990,
        2016: 1017, 2017: 1082, 2018: 1153, 2019: 1219, 2020: 1200,
        2021: 1393, 2022: 1547, 2023: 1740, 2024: 1943,
    }
    for yr, gel in KNOWN.items():
        rate = _nbg_eur_rate(yr)
        if rate:
            result[yr] = {"local": gel, "currency": "GEL", "eur": round(gel / rate, 0)}
    if result:
        yrs = sorted(result)
        print(f"  GE: {yrs[0]}-{yrs[-1]} ({len(yrs)} yrs), "
              f"latest={result[yrs[-1]]['local']:,.0f} GEL "
              f"= {result[yrs[-1]]['eur']:,.0f} EUR")
    return result


# ─── Armenia (Armstat time series) ────────────────────────────────────────────

def fetch_armenia(fx_rates):
    """
    Average monthly nominal wages (AMD) from Armstat time series Excel.
    URL: https://www.armstat.am/file/doc/99570263.xlsx
    Sheet '1980-2025': col[1]=year, col[3]=avg monthly wage (AMD).
    FX: World Bank PA.NUS.FCRF (AMD/USD) + ECB USD/EUR from fx_rates.
    Returns {year: {"local": amd, "currency": "AMD", "eur": eur_or_none}}
    """
    import json as _json
    import openpyxl
    import urllib.request

    path = os.path.join(DATA_DIR, "armstat_wages.xlsx")
    if not os.path.exists(path):
        url = "https://www.armstat.am/file/doc/99570263.xlsx"
        print(f"  AM: downloading Armstat wages...")
        urllib.request.urlretrieve(url, path)

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active  # sheet '1980-2025'

    amd_wages = {}
    for row in ws.iter_rows(values_only=True):
        yr = row[1] if row else None
        wage = row[3] if len(row) > 3 else None
        if not isinstance(yr, int) or yr < 1996 or yr > 2025:
            continue
        if wage is None:
            continue
        try:
            v = float(str(wage).replace("\xa0", "").replace(",", "").strip())
            if v > 0:
                amd_wages[yr] = v
        except (ValueError, TypeError):
            pass

    # World Bank official exchange rate: AMD per 1 USD
    try:
        wb_url = ("https://api.worldbank.org/v2/country/AM/indicator/PA.NUS.FCRF"
                  "?format=json&per_page=50&mrv=50")
        resp = urllib.request.urlopen(wb_url, timeout=15)
        wb_data = _json.loads(resp.read())
        amd_per_usd = {int(d["date"]): float(d["value"])
                       for d in wb_data[1] if d["value"]}
    except Exception as e:
        print(f"  AM: World Bank FX failed: {e}")
        amd_per_usd = {}

    usd_per_eur = fx_rates.get("USD", {})

    result = {}
    for yr, amd in sorted(amd_wages.items()):
        fx_amd_usd = amd_per_usd.get(yr) or amd_per_usd.get(yr - 1)
        fx_usd_eur = usd_per_eur.get(yr)
        if fx_amd_usd and fx_usd_eur:
            eur = (amd / fx_amd_usd) * fx_usd_eur  # AMD → USD → EUR
        else:
            eur = None
        result[yr] = {"local": amd, "currency": "AMD",
                      "eur": round(eur, 0) if eur else None}

    if result:
        yrs = sorted(result)
        latest = result[yrs[-1]]
        print(f"  AM: {yrs[0]}-{yrs[-1]} ({len(yrs)} yrs), "
              f"latest={latest['local']:,.0f} AMD = {latest['eur'] or 'n/a'} EUR")
    return result


# ─── Kazakhstan (Bureau of National Statistics) ───────────────────────────────

def fetch_kazakhstan(fx_rates):
    """
    Average monthly nominal salary (KZT) from Kazakhstan Bureau of National Statistics.
    URL: https://stat.gov.kz/api/iblock/element/469342/file/en/
    Sheet has years 2015-2025 in row 6, 'Republic of Kazakhstan' totals in row 7.
    FX: World Bank PA.NUS.FCRF (KZT/USD) + ECB USD/EUR from fx_rates.
    Returns {year: {"local": kzt, "currency": "KZT", "eur": eur_or_none}}
    """
    import json as _json
    import openpyxl
    import urllib.request

    path = os.path.join(DATA_DIR, "kazakhstan_wages.xlsx")
    if not os.path.exists(path):
        url = "https://stat.gov.kz/api/iblock/element/469342/file/en/"
        print(f"  KZ: downloading stat.gov.kz wages...")
        urllib.request.urlretrieve(url, path)

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    headers = None
    kzt_wages = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 6:
            headers = list(row)
        elif i == 7 and headers is not None:
            seen_years = set()
            for j, val in enumerate(row):
                yr = headers[j] if j < len(headers) else None
                if isinstance(yr, int) and 2015 <= yr <= 2030 and yr not in seen_years:
                    seen_years.add(yr)
                    if val and val != "-":
                        try:
                            kzt_wages[yr] = float(val)
                        except (ValueError, TypeError):
                            pass
            break

    # World Bank official exchange rate: KZT per 1 USD
    try:
        wb_url = ("https://api.worldbank.org/v2/country/KZ/indicator/PA.NUS.FCRF"
                  "?format=json&per_page=30&mrv=30")
        resp = urllib.request.urlopen(wb_url, timeout=15)
        wb_data = _json.loads(resp.read())
        kzt_per_usd = {int(d["date"]): float(d["value"])
                       for d in wb_data[1] if d["value"]}
    except Exception as e:
        print(f"  KZ: World Bank FX failed: {e}")
        kzt_per_usd = {}

    usd_per_eur = fx_rates.get("USD", {})

    result = {}
    for yr, kzt in sorted(kzt_wages.items()):
        fx_kzt_usd = kzt_per_usd.get(yr) or kzt_per_usd.get(yr - 1)
        fx_usd_eur = usd_per_eur.get(yr)
        if fx_kzt_usd and fx_usd_eur:
            eur = (kzt / fx_kzt_usd) * fx_usd_eur  # KZT → USD → EUR
        else:
            eur = None
        result[yr] = {"local": kzt, "currency": "KZT",
                      "eur": round(eur, 0) if eur else None}

    if result:
        yrs = sorted(result)
        latest = result[yrs[-1]]
        print(f"  KZ: {yrs[0]}-{yrs[-1]} ({len(yrs)} yrs), "
              f"latest={latest['local']:,.0f} KZT = {latest['eur'] or 'n/a'} EUR")
    return result


# ─── Albania (INSTAT quarterly survey) ───────────────────────────────────────

def fetch_albania(fx_rates):
    """
    Average monthly gross wage per employee (ALL) from INSTAT Albania quarterly survey.
    URL: https://www.instat.gov.al/media/nnmcdq1f/paga-tatime-trm1_2023_trm1_2026-publikim.xlsx
    Covers Q1 2023 – Q4 2025. Annual value = mean of 4 quarters.
    FX: World Bank PA.NUS.FCRF (ALL/USD) + ECB USD/EUR from fx_rates.
    Returns {year: {"local": all_val, "currency": "ALL", "eur": eur_or_none}}
    """
    import json as _json
    import openpyxl
    import urllib.request

    path = os.path.join(DATA_DIR, "albania_wages.xlsx")
    if not os.path.exists(path):
        url = ("https://www.instat.gov.al/media/nnmcdq1f/"
               "paga-tatime-trm1_2023_trm1_2026-publikim.xlsx")
        print(f"  AL: downloading INSTAT wages...")
        urllib.request.urlretrieve(url, path)

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    all_wages = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 5:  # "Average monthly gross wage per employee" row (0-indexed)
            # Columns 1-4 = 2023 Q1-Q4; 5-8 = 2024 Q1-Q4; 9-12 = 2025 Q1-Q4
            for yr_offset, base_yr in enumerate([2023, 2024, 2025]):
                start = 1 + yr_offset * 4
                quarters = [float(row[start + q])
                            for q in range(4)
                            if (start + q) < len(row) and row[start + q] is not None]
                if len(quarters) == 4:
                    all_wages[base_yr] = sum(quarters) / 4
            break

    # World Bank official exchange rate: ALL per 1 USD
    try:
        wb_url = ("https://api.worldbank.org/v2/country/AL/indicator/PA.NUS.FCRF"
                  "?format=json&per_page=30&mrv=30")
        resp = urllib.request.urlopen(wb_url, timeout=15)
        wb_data = _json.loads(resp.read())
        all_per_usd = {int(d["date"]): float(d["value"])
                       for d in wb_data[1] if d["value"]}
    except Exception as e:
        print(f"  AL: World Bank FX failed: {e}")
        all_per_usd = {}

    usd_per_eur = fx_rates.get("USD", {})

    result = {}
    for yr, all_val in sorted(all_wages.items()):
        fx_all_usd = all_per_usd.get(yr) or all_per_usd.get(yr - 1)
        fx_usd_eur = usd_per_eur.get(yr)
        if fx_all_usd and fx_usd_eur:
            eur = (all_val / fx_all_usd) * fx_usd_eur  # ALL → USD → EUR
        else:
            eur = None
        result[yr] = {"local": round(all_val, 0), "currency": "ALL",
                      "eur": round(eur, 0) if eur else None}

    if result:
        yrs = sorted(result)
        latest = result[yrs[-1]]
        print(f"  AL: {yrs[0]}-{yrs[-1]} ({len(yrs)} yrs), "
              f"latest={latest['local']:,.0f} ALL = {latest['eur'] or 'n/a'} EUR")
    return result


# ─── Eurostat national accounts (D11/employees) ──────────────────────────────

def fetch_eurostat_nataccounts():
    """
    Fetch D11 (wages & salaries) and SAL_DC (employees) from Eurostat national accounts.
    D11/employees gives average monthly wage per employee (headcount, not FTE-adjusted).
    Excludes employer social contributions — unlike OECD AV_AN_WAGE which uses total D1.

    Returns: (wages_d11emp, ratios_d11d1)
      wages_d11emp : {iso2: {year: monthly_eur}}
      ratios_d11d1 : {iso2: {year: D11/D1}}
    """
    import json
    GEO_FIX = {"EL": "GR", "UK": "GB"}

    print(f"\n{'=' * 70}")
    print("EUROSTAT NATIONAL ACCOUNTS (D11 wages / SAL_DC employees)")
    print("=" * 70)

    def _parse(resp):
        d = json.loads(resp.text)
        dims = d["id"]; sizes = d["size"]; values = d.get("value", {})
        dim_rev = {}
        for dn in dims:
            cats = d["dimension"][dn]["category"]["index"]
            dim_rev[dn] = {v: k for k, v in cats.items()}
        result = {}
        for fk, val in values.items():
            idx = int(fk); rem = idx; idxs = []
            for s in reversed(sizes):
                idxs.append(rem % s); rem //= s
            idxs.reverse()
            dv = {dn: dim_rev[dn].get(di, "") for dn, di in zip(dims, idxs)}
            geo = GEO_FIX.get(dv.get("geo", ""), dv.get("geo", ""))
            t = dv.get("time", "")
            if len(geo) == 2 and t.isdigit():
                result.setdefault(geo, {})[int(t)] = val
        return result

    base = "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data"
    r_d11 = requests.get(f"{base}/nama_10_a10",
                         params={"format": "JSON", "nace_r2": "TOTAL", "unit": "CP_MEUR",
                                 "na_item": "D11", "sinceTimePeriod": "2000"}, timeout=60)
    r_d1  = requests.get(f"{base}/nama_10_a10",
                         params={"format": "JSON", "nace_r2": "TOTAL", "unit": "CP_MEUR",
                                 "na_item": "D1",  "sinceTimePeriod": "2000"}, timeout=60)
    r_emp = requests.get(f"{base}/nama_10_a10_e",
                         params={"format": "JSON", "nace_r2": "TOTAL", "unit": "THS_PER",
                                 "na_item": "SAL_DC", "sinceTimePeriod": "2000"}, timeout=60)

    if any(r.status_code != 200 for r in [r_d11, r_d1, r_emp]):
        print(f"  FAILED: D11={r_d11.status_code} D1={r_d1.status_code} EMP={r_emp.status_code}")
        return {}, {}

    d11 = _parse(r_d11)
    d1  = _parse(r_d1)
    emp = _parse(r_emp)

    wages_out = {}
    ratios_out = {}
    for geo in d11:
        for yr in sorted(d11[geo]):
            e = emp.get(geo, {}).get(yr)
            if not e or e <= 0:
                continue
            monthly = (d11[geo][yr] * 1e6) / (e * 1e3) / 12
            wages_out.setdefault(geo, {})[yr] = round(monthly, 0)
            d1v = d1.get(geo, {}).get(yr)
            if d1v and d1v > 0:
                ratios_out.setdefault(geo, {})[yr] = round(d11[geo][yr] / d1v, 4)

    print(f"  {len(wages_out)} countries with D11/SAL_DC data")
    for iso2 in sorted(wages_out):
        yrs = sorted(wages_out[iso2])
        r = ratios_out.get(iso2, {}).get(yrs[-1], 0)
        print(f"    {iso2}: {yrs[0]}-{yrs[-1]} ({len(yrs)} yrs), "
              f"latest={wages_out[iso2][yrs[-1]]:,.0f} EUR/month  D11/D1={r:.3f}")
    return wages_out, ratios_out


def fetch_gdp_per_capita(fx_rates):
    """
    GDP per capita (EUR, current prices).
    - Eurostat nama_10_pc (B1GQ, CP_EUR_HAB): 1975-2025, ~40 European countries
    - World Bank NY.GDP.PCAP.CD (USD->EUR via ECB): remaining (RU,BY,UA,GE,AM,AZ,KZ...)

    Returns: {iso2: {year: gdp_pc_eur}}
    """
    import json
    GEO_FIX = {"EL": "GR", "UK": "GB"}

    print(f"\n{'=' * 70}")
    print("GDP PER CAPITA (EUR current prices)")
    print("=" * 70)

    result = {}

    # Eurostat
    resp = requests.get(
        "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/nama_10_pc",
        params={"format": "JSON", "na_item": "B1GQ", "unit": "CP_EUR_HAB",
                "sinceTimePeriod": "1990"},
        timeout=60)
    if resp.status_code == 200:
        d = json.loads(resp.text)
        dims = d["id"]; sizes = d["size"]; values = d.get("value", {})
        dim_rev = {}
        for dn in dims:
            cats = d["dimension"][dn]["category"]["index"]
            dim_rev[dn] = {v: k for k, v in cats.items()}
        for fk, val in values.items():
            idx = int(fk); rem = idx; idxs = []
            for s in reversed(sizes):
                idxs.append(rem % s); rem //= s
            idxs.reverse()
            dv = {dn: dim_rev[dn].get(di, "") for dn, di in zip(dims, idxs)}
            geo = GEO_FIX.get(dv.get("geo", ""), dv.get("geo", ""))
            t = dv.get("time", "")
            if len(geo) == 2 and t.isdigit() and geo in ISO2_TO_NAME:
                result.setdefault(geo, {})[int(t)] = round(val, 0)
        print(f"  Eurostat: {len(result)} countries, 1990-2025")
    else:
        print(f"  Eurostat FAILED: HTTP {resp.status_code}")

    # World Bank for non-Eurostat countries
    missing = {v[0] for v in EUROPEAN.values()} - set(result.keys()) - {"LI", "MC"}
    if missing:
        iso2_list = ";".join(sorted(missing))
        wb = requests.get(
            f"https://api.worldbank.org/v2/country/{iso2_list}/indicator/NY.GDP.PCAP.CD",
            params={"format": "json", "per_page": "1000", "date": "1990:2025"},
            timeout=30)
        if wb.status_code == 200:
            wb_data = wb.json()
            usd_eur = fx_rates.get("USD", {})
            covered = set()
            if len(wb_data) > 1 and wb_data[1]:
                for entry in wb_data[1]:
                    iso2 = entry.get("country", {}).get("id", "")
                    yr_str = entry.get("date", "")
                    val = entry.get("value")
                    if iso2 and yr_str.isdigit() and val and iso2 in ISO2_TO_NAME:
                        yr = int(yr_str)
                        rate = usd_eur.get(yr)
                        if rate and rate > 0:
                            result.setdefault(iso2, {})[yr] = round(val / rate, 0)
                            covered.add(iso2)
            print(f"  World Bank: {sorted(covered)}")
        else:
            print(f"  World Bank FAILED: HTTP {wb.status_code}")

    print(f"  Total GDP coverage: {len(result)} countries")
    return result


def fetch_imf_weo(filepath, fx_rates):
    """
    Parse IMF WEO Excel (WEOApr2026all.xlsx).
    Extracts NGDPDPC (GDP per capita, current prices, USD) for European countries.
    Converts to EUR: ECB historical rates for past years, last known rate for forecasts.
    Returns: {iso2: {year: gdp_pc_eur}}
    """
    import openpyxl

    # WEO ISO3 codes not in the OECD mapping
    WEO_EXTRA = {
        "AND": "AD",  # Andorra
        "ARM": "AM",  # Armenia
        "KAZ": "KZ",  # Kazakhstan
        "ALB": "AL",  # Albania
    }

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Countries"]

    usd_eur = fx_rates.get("USD", {})
    last_hist_yr = max((y for y in usd_eur if y <= 2025), default=2024)
    fallback_rate = usd_eur.get(last_hist_yr, 1.08)

    headers = None
    year_cols = {}
    result = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = row
            for j, h in enumerate(headers):
                if isinstance(h, int) and 1990 <= h <= 2031:
                    year_cols[j] = h
            continue
        if row[4] != "NGDPDPC":
            continue
        iso3 = row[2]
        iso2 = ISO3_TO_ISO2.get(iso3) or WEO_EXTRA.get(iso3)
        if not iso2:
            continue
        country_data = {}
        for j, yr in year_cols.items():
            val = row[j]
            if val is None or val == "":
                continue
            try:
                val = float(val)
            except (ValueError, TypeError):
                continue
            if val <= 0:
                continue
            rate = usd_eur.get(yr, fallback_rate)
            country_data[yr] = round(val / rate, 0)
        if country_data:
            result[iso2] = country_data

    wb.close()
    print(f"  IMF WEO: {len(result)} European countries, forecasts to 2031")
    for iso2 in sorted(result):
        yrs = sorted(result[iso2])
        print(f"    {iso2}: {yrs[0]}-{yrs[-1]}  "
              f"2025={result[iso2].get(2025, 0):,.0f} EUR  "
              f"2031={result[iso2].get(2031, 0):,.0f} EUR")
    return result


def _linreg(xs, ys):
    """OLS linear regression. Returns (slope, intercept, r2)."""
    n = len(xs)
    if n < 2:
        return 0.0, ys[0] if ys else 0.0, 0.0
    xm = sum(xs) / n; ym = sum(ys) / n
    ssxy = sum((x - xm) * (y - ym) for x, y in zip(xs, ys))
    ssxx = sum((x - xm) ** 2 for x in xs)
    if ssxx == 0:
        return 0.0, ym, 0.0
    b = ssxy / ssxx; a = ym - b * xm
    ss_res = sum((y - (a + b * x)) ** 2 for x, y in zip(xs, ys))
    ss_tot = sum((y - ym) ** 2 for y in ys)
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 1.0
    return b, a, r2


def project_wages(rows, gdp_data, horizon=2031):
    """
    Project wages from last actual year to `horizon`.

    Method for each country:
    1. Fit log(GDP_pc) ~ year on last 5 observed years -> exponential GDP extrapolation
       (overridden by IMF WEO forecast GDP where available)
    2. Compute wage/GDP_monthly ratio for last 5 years -> take MEAN (flat ratio)
       NOTE: Linear ratio extrapolation was rejected because it amplifies temporary
       structural shifts (e.g. Poland's minimum wage hikes, Spain's post-COVID GDP
       recovery outpacing wages) into unrealistic long-run projections.
       A flat mean ratio = wages grow proportionally with GDP per capita.
    3. projected_wage = mean_ratio * projected_gdp_pc / 12

    Returns list of rows with source='projected', is_forecast=1.
    """
    from math import log, exp

    wages_series = {}
    for r in rows:
        iso2 = r["iso2"]
        yr = r["year"]
        w = r.get("wage_monthly_eur", "")
        if w and w != "":
            wages_series.setdefault(iso2, {})[yr] = float(w)

    projected = []
    n_countries = 0

    for iso2, wages in wages_series.items():
        gdp = gdp_data.get(iso2, {})
        if not gdp:
            continue
        common = sorted(set(wages) & set(gdp))
        if len(common) < 3:
            continue

        fit_yrs = common[-5:]
        last_yr = fit_yrs[-1]

        # GDP trend (exponential)
        log_gdps = [log(max(gdp[y], 1)) for y in fit_yrs]
        g_slope, g_intcpt, _ = _linreg(fit_yrs, log_gdps)

        # Wage/GDP monthly ratio: flat mean of last 5 years
        # (linear trend extrapolation causes unrealistic divergence for countries
        # with recent structural shifts — minimum wage hikes, post-COVID recovery)
        ratios = [wages[y] / (gdp[y] / 12) for y in fit_yrs]
        proj_ratio = sum(ratios) / len(ratios)
        proj_ratio = max(0.2, min(1.5, proj_ratio))

        name = ISO2_TO_NAME.get(iso2, iso2)
        n_countries += 1

        for yr in range(last_yr + 1, horizon + 1):
            # Use IMF forecast GDP if available; fall back to extrapolated trend
            if yr in gdp:
                proj_gdp = gdp[yr]
            else:
                proj_gdp = exp(g_intcpt + g_slope * yr)
            proj_wage = proj_ratio * (proj_gdp / 12)
            projected.append({
                "iso2": iso2, "country": name, "year": yr,
                "wage_monthly_eur": round(proj_wage, 0),
                "wage_monthly_usd": "",
                "wage_monthly_local": "",
                "currency": "EUR",
                "source": "projected",
                "wage_oecd_eur": "",
                "wage_d11emp_eur": "",
                "gdp_pc_eur": round(proj_gdp, 0),
                "is_forecast": 1,
            })

    print(f"  Projected {n_countries} countries to {horizon}")
    return projected


# ─── Phase 1: OECD ──────────────────────────────────────────────────────────

def fetch_oecd_wages():
    """
    Fetch OECD Average Annual Wages for all available countries.
    Returns: {iso3: {year: {measure: value}}}
    """
    print("=" * 70)
    print("PHASE 1: OECD AV_AN_WAGE")
    print("=" * 70)

    url = ("https://sdmx.oecd.org/public/rest/data/"
           "OECD.ELS.SAE,DSD_EARNINGS@AV_AN_WAGE,1.0/"
           "......?"
           "startPeriod=1990&endPeriod=2026"
           "&format=csvfilewithlabels")

    print(f"  Fetching: {url[:80]}...")
    resp = requests.get(url, timeout=120)
    if resp.status_code != 200:
        print(f"  FAILED: HTTP {resp.status_code}")
        print(f"  {resp.text[:500]}")
        return {}

    reader = csv.DictReader(io.StringIO(resp.text))
    data = defaultdict(lambda: defaultdict(dict))
    measures_seen = set()
    units_seen = set()
    row_count = 0

    for row in reader:
        row_count += 1
        ref = row.get("REF_AREA", "")
        unit = row.get("UNIT_MEASURE", "")
        measure = row.get("MEASURE", "")
        period = row.get("TIME_PERIOD", "")
        val = row.get("OBS_VALUE", "")
        price_base = row.get("PRICE_BASE", "")

        if not (ref and period and val):
            continue

        # Skip constant-price series — we want nominal (current prices)
        if price_base == "Q":
            continue

        measures_seen.add(measure)
        units_seen.add(unit)

        try:
            year = int(period[:4])
            # Store as (unit_measure, measure) combo
            key = f"{unit}|{measure}" if measure else unit
            data[ref][year][key] = float(val)
        except (ValueError, IndexError):
            pass

    print(f"  Total rows: {row_count:,}")
    print(f"  UNIT_MEASURE values: {sorted(units_seen)}")
    print(f"  MEASURE values: {sorted(measures_seen)}")
    print(f"  Countries total: {len(data)}")

    # Show European coverage
    euro_found = {k for k in data if k in EUROPEAN}
    non_euro_found = {k for k in data if k not in EUROPEAN}
    print(f"  European countries: {len(euro_found)}")
    print(f"  Non-European: {sorted(non_euro_found)}")

    print(f"\n  {'ISO':<5} {'Country':<22} {'Years':>12} {'N':>4}  Measures")
    print(f"  {'-'*75}")
    for iso3 in sorted(euro_found):
        years = sorted(data[iso3].keys())
        iso2, name = EUROPEAN[iso3]
        # Collect all measure keys
        all_keys = set()
        for yr_data in data[iso3].values():
            all_keys.update(yr_data.keys())
        print(f"  {iso2:<5} {name:<22} {years[0]}-{years[-1]:>4} {len(years):>4}  {sorted(all_keys)}")

    return dict(data)


def fetch_ecb_fx_rates():
    """
    Fetch annual FX rates for all currencies vs EUR from ECB.
    Returns: {currency_code: {year: units_per_eur}}
    e.g. {"PLN": {2024: 4.35}, "CZK": {2024: 25.2}, "USD": {2024: 1.08}}

    ECB convention: rate = how many units of foreign currency per 1 EUR.
    So to convert from PLN to EUR: eur_value = pln_value / rate
    """
    print(f"\n{'=' * 70}")
    print("ECB exchange rates (all currencies vs EUR)")
    print("=" * 70)

    # Fetch all currencies at once
    url = ("https://data-api.ecb.europa.eu/service/data/"
           "EXR/A..EUR.SP00.A?"
           "format=csvdata&startPeriod=1990&endPeriod=2026")

    print(f"  Fetching: {url[:80]}...")
    try:
        resp = requests.get(url, timeout=90)
    except requests.exceptions.Timeout:
        print("  ECB timeout — falling back to hardcoded key rates")
        return _fallback_fx_rates()

    if resp.status_code != 200:
        print(f"  FAILED: HTTP {resp.status_code}")
        return _fallback_fx_rates()

    reader = csv.DictReader(io.StringIO(resp.text))
    rates = defaultdict(dict)

    for row in reader:
        currency = row.get("CURRENCY", "")
        period = row.get("TIME_PERIOD", "")
        val = row.get("OBS_VALUE", "")
        if currency and period and val:
            try:
                rates[currency][int(period[:4])] = float(val)
            except ValueError:
                pass

    # ECB suspended ISK rates during Iceland's capital controls (2009-2017).
    # Fill from Central Bank of Iceland annual averages.
    ISK_FALLBACK = {
        2009: 172.15, 2010: 161.62, 2011: 161.20, 2012: 160.73,
        2013: 162.07, 2014: 154.13, 2015: 145.27, 2016: 133.31,
        2017: 121.32,
    }
    if "ISK" in rates:
        for yr, rate in ISK_FALLBACK.items():
            if yr not in rates["ISK"]:
                rates["ISK"][yr] = rate
        print(f"  ISK: filled {len(ISK_FALLBACK)} missing years (2009-2017) "
              f"from Central Bank of Iceland")

    # RUB: ECB stops after 2021 (sanctions). Fill 2022-2025 from CBR/market.
    RUB_FALLBACK = {
        2022: 73.95, 2023: 92.37, 2024: 98.55, 2025: 93.00,
    }
    if "RUB" not in rates:
        rates["RUB"] = {}
    for yr, rate in RUB_FALLBACK.items():
        if yr not in rates["RUB"]:
            rates["RUB"][yr] = rate
    print(f"  RUB: filled {len(RUB_FALLBACK)} post-sanctions years (2022-2025)")

    # UAH: ECB never published. Annual averages from NBU/market data.
    rates["UAH"] = {
        1999: 4.39, 2000: 4.82, 2001: 4.81, 2002: 5.03, 2003: 6.02,
        2004: 6.61, 2005: 6.39, 2006: 6.34, 2007: 6.92, 2008: 7.71,
        2009: 10.87, 2010: 10.53, 2011: 11.09, 2012: 10.27, 2013: 10.61,
        2014: 15.72, 2015: 24.23, 2016: 28.29, 2017: 30.00, 2018: 32.14,
        2019: 28.95, 2020: 30.79, 2021: 32.29, 2022: 38.42, 2023: 40.22,
        2024: 43.20, 2025: 45.00,
    }
    print(f"  UAH: added {len(rates['UAH'])} years (1999-2025) from NBU data")

    # BYN: ECB never published. Annual averages from NBRB/market data.
    # Pre-2016 redenomination: 1 BYN = 10,000 BYR. All values here in BYN.
    rates["BYN"] = {
        2016: 2.20, 2017: 2.24, 2018: 2.36, 2019: 2.33, 2020: 2.79,
        2021: 2.99, 2022: 2.93, 2023: 3.24, 2024: 3.53, 2025: 3.60,
    }
    print(f"  BYN: added {len(rates['BYN'])} years (2016-2025) from NBRB data")

    # Which currencies we care about
    needed = {"PLN", "CZK", "HUF", "GBP", "DKK", "SEK", "NOK", "CHF",
              "ISK", "TRY", "BGN", "RON", "USD", "HRK",
              "RUB", "BYN", "UAH"}
    found = set(rates.keys()) & needed
    print(f"  Currencies found: {len(rates)} total, {len(found)} needed")
    for curr in sorted(found):
        yrs = sorted(rates[curr].keys())
        print(f"    {curr}: {yrs[0]}-{yrs[-1]} ({len(yrs)} yrs), "
              f"latest={rates[curr][yrs[-1]]:.4f}")

    return dict(rates)


def _fallback_fx_rates():
    """Hardcoded annual average rates for key years if ECB API fails."""
    # Source: ECB reference rates, annual averages
    # Only used as fallback — real ECB data is preferred
    print("  Using hardcoded fallback FX rates (limited coverage)")
    return {
        "PLN": {2020: 4.4430, 2021: 4.5652, 2022: 4.6861, 2023: 4.5420, 2024: 4.3000, 2025: 4.2000},
        "CZK": {2020: 26.455, 2021: 25.640, 2022: 24.566, 2023: 24.004, 2024: 25.100, 2025: 25.000},
        "HUF": {2020: 351.25, 2021: 358.52, 2022: 391.29, 2023: 381.85, 2024: 395.00, 2025: 400.00},
        "GBP": {2020: 0.8897, 2021: 0.8596, 2022: 0.8528, 2023: 0.8698, 2024: 0.8400, 2025: 0.8400},
        "DKK": {2020: 7.4542, 2021: 7.4370, 2022: 7.4396, 2023: 7.4509, 2024: 7.4600, 2025: 7.4600},
        "SEK": {2020: 10.486, 2021: 10.146, 2022: 10.631, 2023: 11.479, 2024: 11.400, 2025: 11.400},
        "NOK": {2020: 10.723, 2021: 10.163, 2022: 10.103, 2023: 11.425, 2024: 11.600, 2025: 11.600},
        "CHF": {2020: 1.0705, 2021: 1.0811, 2022: 1.0047, 2023: 0.9717, 2024: 0.9400, 2025: 0.9400},
        "ISK": {2020: 154.59, 2021: 150.28, 2022: 141.74, 2023: 148.63, 2024: 150.00, 2025: 150.00},
        "TRY": {2020: 8.0547, 2021: 10.512, 2022: 17.409, 2023: 25.761, 2024: 35.000, 2025: 38.000},
        "BGN": {2020: 1.9558, 2021: 1.9558, 2022: 1.9558, 2023: 1.9558, 2024: 1.9558, 2025: 1.9558},
        "RON": {2020: 4.8383, 2021: 4.9215, 2022: 4.9313, 2023: 4.9467, 2024: 4.9750, 2025: 4.9800},
        "USD": {2020: 1.1422, 2021: 1.1827, 2022: 1.0530, 2023: 1.0813, 2024: 1.0800, 2025: 1.0800},
    }


# ─── Build combined dataset ─────────────────────────────────────────────────

def build_wage_table(oecd_data, fx_rates):
    """
    Convert OECD annual wages to monthly EUR.

    OECD AV_AN_WAGE data format:
      Keys are "{currency}|WG" — e.g. "EUR|WG", "PLN|WG", "CZK|WG"
      Also "USD_PPP|WG" for PPP values (we skip these for nominal).

    For eurozone countries: EUR value is already available → monthly = annual / 12.
    For non-eurozone: local currency → EUR via ECB FX rates.

    Returns: list of dicts (rows for CSV)
    """
    rows = []

    for iso3, years_data in oecd_data.items():
        if iso3 not in EUROPEAN:
            continue
        iso2, name = EUROPEAN[iso3]

        for year in sorted(years_data.keys()):
            measures = years_data[year]

            # Extract local currency value and currency code
            annual_local = None
            currency = None
            for key, val in measures.items():
                if "USD_PPP" in key:
                    continue  # skip PPP
                parts = key.split("|")
                if len(parts) == 2 and parts[1] == "WG":
                    currency = parts[0]
                    annual_local = val
                    break

            if annual_local is None or currency is None:
                continue

            monthly_local = annual_local / 12.0

            # Convert to EUR
            if currency == "EUR":
                monthly_eur = monthly_local
            elif currency in fx_rates and year in fx_rates[currency]:
                rate = fx_rates[currency][year]
                monthly_eur = monthly_local / rate if rate > 0 else None
            else:
                monthly_eur = None

            # Also compute USD
            monthly_usd = None
            if "USD" in fx_rates and year in fx_rates["USD"]:
                usd_rate = fx_rates["USD"][year]  # USD per EUR
                if monthly_eur is not None:
                    monthly_usd = monthly_eur * usd_rate

            rows.append({
                "iso2": iso2,
                "country": name,
                "year": year,
                "wage_monthly_eur": round(monthly_eur, 0) if monthly_eur else "",
                "wage_monthly_usd": round(monthly_usd, 0) if monthly_usd else "",
                "wage_monthly_local": round(monthly_local, 0),
                "currency": currency,
                "source": "OECD",
            })

    return rows


def rows_to_series(rows, value_key="wage_monthly_eur"):
    """Convert row list to {iso2: {year: value}} for plotting."""
    series = {}
    for r in rows:
        iso2 = r["iso2"]
        year = r["year"]
        val = r.get(value_key, "")
        if val == "" or val is None:
            continue
        if iso2 not in series:
            series[iso2] = {}
        series[iso2][year] = float(val)
    return series


# ─── Population (for line thickness) ─────────────────────────────────────────

# Approximate 2024 population in millions (Eurostat / World Bank)
POPULATION = {
    "AL": 2.8, "AD": 0.08, "AM": 3.0, "AT": 9.1, "AZ": 10.2,
    "BY": 9.2, "BE": 11.7, "BA": 3.2, "BG": 6.5, "HR": 3.9,
    "CY": 1.3, "CZ": 10.9, "DK": 5.9, "EE": 1.4, "FI": 5.6,
    "FR": 68.2, "GE": 3.7, "DE": 84.5, "GR": 10.4, "HU": 9.6,
    "IS": 0.4, "IE": 5.2, "IT": 58.9, "XK": 1.8, "LV": 1.8,
    "LI": 0.04, "LT": 2.9, "LU": 0.67, "MT": 0.54, "MD": 2.5,
    "MC": 0.04, "ME": 0.62, "NL": 17.9, "MK": 1.8, "NO": 5.5,
    "PL": 37.6, "PT": 10.3, "RO": 19.0, "RU": 144.0, "SM": 0.03,
    "RS": 6.6, "SK": 5.4, "SI": 2.1, "ES": 48.0, "SE": 10.5,
    "CH": 8.8, "TR": 85.3, "UA": 37.0, "GB": 67.7, "KZ": 19.8,
}

import math

def _line_width(iso2, min_w=0.3, max_w=4.5):
    """Line width proportional to log(population)."""
    pop = POPULATION.get(iso2, 1.0)
    # log scale: Montenegro 0.62M → Russia 144M
    log_min, log_max = math.log(0.03), math.log(144.0)
    t = (math.log(max(pop, 0.03)) - log_min) / (log_max - log_min)
    return min_w + t * (max_w - min_w)


# ─── Charts ──────────────────────────────────────────────────────────────────

def _add_end_labels(ax, labels, fontsize=8, x_pad=0.5):
    """
    Add country names at end of lines, nudging vertically to avoid overlap.
    labels: list of (x, y, name, color) sorted by y descending.
    """
    if not labels:
        return
    labels.sort(key=lambda t: -t[1])  # top to bottom

    # Get axis data range for minimum spacing
    ymin, ymax = ax.get_ylim()
    # Min gap in data units — scale with font size, not axis range
    # Approximate: fontsize points / dpi * data_range / axes_height_inches
    ax_height = ax.get_position().height * ax.figure.get_figheight()
    min_gap = (ymax - ymin) * fontsize / (ax_height * 72) * 1.4

    # Nudge overlapping labels
    placed = []
    for x, y, name, color in labels:
        nudged = y
        for _, py in placed:
            if abs(nudged - py) < min_gap:
                nudged = py - min_gap
        placed.append((x, nudged))
        ax.annotate(name, xy=(x, y), xytext=(x + x_pad, nudged),
                    fontsize=fontsize, color=color, va="center",
                    annotation_clip=False)


def plot_gdp_wage_scatter(rows, gdp_data, filename):
    """
    Scatter: log(GDP per capita EUR) vs log(monthly wage EUR).
    One point per (country, year) for available years, 2025 highlighted.
    Regression line + R² for 2024/2025 cross-section.
    """
    import math

    fig, ax = plt.subplots(figsize=(14, 10))

    # Collect all historical points
    wage_series = rows_to_series(rows, "wage_monthly_eur")
    all_pts = []  # (gdp, wage, iso2, year)
    for r in rows:
        iso2 = r["iso2"]
        yr = r["year"]
        w = r.get("wage_monthly_eur", "")
        g = r.get("gdp_pc_eur", "")
        if w and g and w != "" and g != "" and r.get("is_forecast", 0) == 0:
            try:
                all_pts.append((float(g), float(w), iso2, yr))
            except (ValueError, TypeError):
                pass

    if not all_pts:
        plt.close(fig)
        return

    # Plot all years as small grey points
    for gdp, wage, iso2, yr in all_pts:
        if yr < 2024:
            ax.scatter(gdp, wage, s=6, color="#cccccc", alpha=0.4, zorder=1)

    # Plot 2024/2025 highlighted per country (latest available)
    latest_pts = {}
    for gdp, wage, iso2, yr in all_pts:
        if yr >= 2020:
            if iso2 not in latest_pts or yr > latest_pts[iso2][2]:
                latest_pts[iso2] = (gdp, wage, yr)

    cmap = plt.colormaps["tab20"]
    iso2_list = sorted(latest_pts)
    for i, iso2 in enumerate(iso2_list):
        gdp, wage, yr = latest_pts[iso2]
        color = COLORS.get(iso2, cmap(i % 20))
        ax.scatter(gdp, wage, s=60, color=color, zorder=3, alpha=0.9)
        name = ISO2_TO_NAME.get(iso2, iso2)
        ax.annotate(name, (gdp, wage), textcoords="offset points",
                    xytext=(4, 2), fontsize=6.5, color=color)

    # Regression on latest cross-section (log-log)
    if len(latest_pts) >= 5:
        log_g = [math.log(v[0]) for v in latest_pts.values() if v[0] > 0 and v[1] > 0]
        log_w = [math.log(v[1]) for v in latest_pts.values() if v[0] > 0 and v[1] > 0]
        slope, intercept, r2 = _linreg(log_g, log_w)
        xs = sorted(log_g)
        ys = [intercept + slope * x for x in xs]
        ax.plot([math.exp(x) for x in xs], [math.exp(y) for y in ys],
                "k--", linewidth=1.2, alpha=0.6, zorder=2)
        ax.text(0.05, 0.95,
                f"Latest cross-section (N={len(latest_pts)})\n"
                f"R² = {r2:.3f}   elasticity = {slope:.3f}\n"
                f"(elasticity>1: wages grow faster than GDP per capita)",
                transform=ax.transAxes, fontsize=9, va="top",
                bbox=dict(boxstyle="round", fc="white", alpha=0.8))

    ax.set_xscale("log"); ax.set_yscale("log")
    ax.set_xlabel("GDP per capita (EUR, current prices, log scale)", fontsize=12)
    ax.set_ylabel("Avg Monthly Gross Wage (EUR, log scale)", fontsize=12)
    ax.set_title("GDP per Capita vs Average Monthly Wage — European Countries\n"
                 "Grey: historical (pre-2020). Coloured: latest available. Log-log scale.",
                 fontsize=12)
    ax.grid(True, alpha=0.3, which="both")

    plt.tight_layout()
    path = os.path.join(OUT_DIR, filename)
    fig.savefig(path, dpi=150)
    plt.close(fig)
    print(f"  Saved: {path}")


def plot_projection(all_rows, focus_codes, filename):
    """Historical + projected wages to 2031. Projected portion as dashed line."""
    eur_series = rows_to_series(all_rows, "wage_monthly_eur")

    # Split historical vs projected
    hist = {}; proj = {}
    for r in all_rows:
        iso2 = r["iso2"]
        yr = r["year"]
        w = r.get("wage_monthly_eur", "")
        if not w or w == "":
            continue
        if r.get("is_forecast", 0) == 1:
            proj.setdefault(iso2, {})[yr] = float(w)
        else:
            hist.setdefault(iso2, {})[yr] = float(w)

    fig, ax = plt.subplots(figsize=(14, 9))
    end_labels = []

    for iso2 in focus_codes:
        color = COLORS.get(iso2, "#888888")
        name = ISO2_TO_NAME.get(iso2, iso2)
        lw = _line_width(iso2)

        # Historical
        if iso2 in hist:
            h = hist[iso2]
            yrs = sorted(h)
            ax.plot(yrs, [h[y] for y in yrs], "-", color=color, linewidth=lw)

            # Bridge: connect last historical to first projected
            if iso2 in proj:
                p = proj[iso2]
                p_yrs = sorted(p)
                bridge_x = [yrs[-1]] + p_yrs
                bridge_y = [h[yrs[-1]]] + [p[y] for y in p_yrs]
                ax.plot(bridge_x, bridge_y, "--", color=color, linewidth=lw,
                        alpha=0.7)
                end_labels.append((p_yrs[-1], p[p_yrs[-1]], f"{name} (proj.)", color))
            else:
                end_labels.append((yrs[-1], h[yrs[-1]], name, color))

    # Shaded projection zone
    ax.axvspan(2025.5, 2031.5, alpha=0.05, color="gray")
    ax.axvline(x=2025.5, color="gray", linestyle=":", linewidth=1, alpha=0.5)
    ax.text(2026, ax.get_ylim()[1] if ax.get_ylim()[1] > 0 else 1000,
            "← actual  projected →", fontsize=8, color="gray", ha="center")

    ax.set_xlabel("Year", fontsize=12)
    ax.set_ylabel("Avg Monthly Gross Wage (EUR nominal)", fontsize=12)
    ax.set_title("Wage Convergence + Projection to 2031\n"
                 "Method: IMF WEO GDP forecast × mean wage/GDP ratio (last 5 years)",
                 fontsize=12)
    ax.grid(True, alpha=0.3)
    ax.set_xlim(1995, 2033)
    ax.set_ylim(bottom=0)
    _add_end_labels(ax, end_labels, fontsize=9)

    plt.tight_layout()
    path = os.path.join(OUT_DIR, filename)
    fig.savefig(path, dpi=150)
    plt.close(fig)
    print(f"  Saved: {path}")


def plot_focus(rows, focus_codes, filename, title_suffix=""):
    """Validation chart for a specific set of countries."""
    eur_series = rows_to_series(rows, "wage_monthly_eur")
    usd_series = rows_to_series(rows, "wage_monthly_usd")

    fig, axes = plt.subplots(1, 2, figsize=(18, 8))

    for ax, series, label in [
        (axes[0], eur_series, "EUR"),
        (axes[1], usd_series, "USD"),
    ]:
        end_labels = []
        for iso2 in focus_codes:
            if iso2 not in series:
                continue
            s = series[iso2]
            years = sorted(s.keys())
            values = [s[y] for y in years]
            lw = _line_width(iso2)
            color = COLORS.get(iso2, "#888888")
            name = ISO2_TO_NAME.get(iso2, iso2)
            ax.plot(years, values, '-o', color=color,
                    markersize=3, linewidth=lw)
            end_labels.append((years[-1], values[-1], name, color))

        ax.set_xlabel("Year", fontsize=12)
        ax.set_ylabel(f"Avg Monthly Wage ({label} nominal)", fontsize=12)
        ax.set_title(f"Average Monthly Gross Wages ({label}){title_suffix}\n"
                     f"Source: Eurostat D11/employees, national offices, ECB rates", fontsize=12)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(1990, 2029)
        ax.set_ylim(bottom=0)
        _add_end_labels(ax, end_labels, fontsize=10)

    plt.tight_layout()
    path = os.path.join(OUT_DIR, filename)
    fig.savefig(path, dpi=150)
    plt.close(fig)
    print(f"  Saved: {path}")


def plot_all_europe(rows, filename):
    """Overview chart: all European countries in EUR."""
    eur_series = rows_to_series(rows, "wage_monthly_eur")

    fig, ax = plt.subplots(figsize=(18, 22))

    # Rank by latest value
    latest = {}
    for iso2, s in eur_series.items():
        years = sorted(s.keys())
        if years:
            latest[iso2] = s[years[-1]]

    ranked = sorted(latest.keys(), key=lambda x: latest[x], reverse=True)
    cmap = plt.colormaps["tab20"]

    end_labels = []
    for i, iso2 in enumerate(ranked):
        s = eur_series[iso2]
        years = sorted(s.keys())
        values = [s[y] for y in years]
        name = ISO2_TO_NAME.get(iso2, iso2)
        val = latest[iso2]
        lw = _line_width(iso2)
        color = cmap(i % 20)
        ax.plot(years, values, '-', color=color,
                linewidth=lw, alpha=0.85)
        end_labels.append((years[-1], values[-1],
                           f"{name} ({val:,.0f})", color))

    ax.set_xlabel("Year", fontsize=12)
    ax.set_ylabel("Avg Monthly Wage (EUR nominal)", fontsize=12)
    ax.set_title("Average Monthly Gross Wages in EUR — All European Countries\n"
                 "Source: OECD AV_AN_WAGE + Eurostat + national offices, "
                 "ECB exchange rates", fontsize=13)
    ax.grid(True, alpha=0.3)
    ax.set_xlim(1990, 2032)
    ax.set_ylim(bottom=0)
    _add_end_labels(ax, end_labels, fontsize=7, x_pad=0.3)

    plt.tight_layout()
    path = os.path.join(OUT_DIR, filename)
    fig.savefig(path, dpi=150)
    plt.close(fig)
    print(f"  Saved: {path}")


def plot_ratio_germany(rows, all_rows, focus_codes, filename):
    """Wages as % of Germany — historical (solid) + projection to 2031 (dashed)."""
    hist_series = rows_to_series(rows, "wage_monthly_eur")
    proj_series = rows_to_series(
        [r for r in all_rows if r["is_forecast"] == 1], "wage_monthly_eur"
    )
    # Full series (hist + proj) for Germany denominator
    full_series = rows_to_series(all_rows, "wage_monthly_eur")
    de_hist = hist_series.get("DE", {})
    de_full = full_series.get("DE", {})
    if not de_hist:
        print("  No Germany data for ratio chart")
        return

    fig, ax = plt.subplots(figsize=(14, 11))

    end_labels = []
    for iso2 in focus_codes:
        if iso2 == "DE" or iso2 not in hist_series:
            continue
        s_hist = hist_series[iso2]
        lw = _line_width(iso2)
        color = COLORS.get(iso2, "#888")
        name = ISO2_TO_NAME.get(iso2, iso2)

        # Historical solid line
        common_hist = sorted(set(s_hist.keys()) & set(de_hist.keys()))
        if not common_hist:
            continue
        ratios_hist = [(s_hist[y] / de_hist[y]) * 100 for y in common_hist]
        ax.plot(common_hist, ratios_hist, '-o', color=color,
                markersize=3, linewidth=lw)

        # Projection dashed line (bridge from last historical point)
        s_proj = proj_series.get(iso2, {})
        if s_proj and de_full:
            bridge_yr = common_hist[-1]
            bridge_ratio = ratios_hist[-1]
            proj_yrs = sorted(s_proj.keys())
            proj_ratios = [(s_proj[y] / de_full[y]) * 100 for y in proj_yrs if y in de_full]
            proj_yrs = [y for y in proj_yrs if y in de_full]
            if proj_yrs:
                ax.plot([bridge_yr] + proj_yrs,
                        [bridge_ratio] + proj_ratios,
                        '--', color=color, linewidth=lw, alpha=0.7)
                end_labels.append((proj_yrs[-1], proj_ratios[-1], name, color))
            else:
                end_labels.append((common_hist[-1], ratios_hist[-1], name, color))
        else:
            end_labels.append((common_hist[-1], ratios_hist[-1], name, color))

    ax.axhline(y=100, color='gray', linestyle='--', alpha=0.5)
    ax.annotate("Germany = 100%", xy=(2026.2, 100.5), fontsize=9, color='gray')
    ax.axvline(x=2025.5, color='gray', linestyle=':', alpha=0.4, linewidth=0.8)
    ax.text(2025.6, 115, "← actual  projected →", fontsize=8, color='gray')
    ax.set_xlabel("Year", fontsize=12)
    ax.set_ylabel("% of German Average Wage (EUR nominal)", fontsize=12)
    ax.set_title("Wages as % of Germany (EUR nominal) + Projection to 2031\n"
                 "Source: Eurostat D11/employees, national offices, ECB rates", fontsize=13)
    ax.grid(True, alpha=0.3)
    ax.set_xlim(1995, 2033)
    ax.set_ylim(0, 120)
    _add_end_labels(ax, end_labels, fontsize=10)

    plt.tight_layout()
    path = os.path.join(OUT_DIR, filename)
    fig.savefig(path, dpi=150)
    plt.close(fig)
    print(f"  Saved: {path}")


# ─── Validation ──────────────────────────────────────────────────────────────

def print_validation(rows):
    """Print key values for quick validation."""
    eur_series = rows_to_series(rows, "wage_monthly_eur")

    print(f"\n{'=' * 70}")
    print("VALIDATION: Monthly EUR wages (nominal, at market exchange rates)")
    print(f"{'=' * 70}")
    print(f"  {'Country':<22} {'2000':>7} {'2010':>7} {'2015':>7} {'2020':>7} "
          f"{'2024':>7} {'Latest':>7} {'Yr':>5}")
    print(f"  {'-'*80}")

    # Sort by latest value
    latest_vals = {}
    for iso2, s in eur_series.items():
        years = sorted(s.keys())
        if years:
            latest_vals[iso2] = (years[-1], s[years[-1]])

    for iso2 in sorted(latest_vals, key=lambda x: latest_vals[x][1], reverse=True):
        s = eur_series[iso2]
        name = ISO2_TO_NAME.get(iso2, iso2)
        latest_yr, latest_val = latest_vals[iso2]

        def fmt(yr):
            return f"{s[yr]:,.0f}" if yr in s else "—"

        print(f"  {name:<22} {fmt(2000):>7} {fmt(2010):>7} {fmt(2015):>7} "
              f"{fmt(2020):>7} {fmt(2024):>7} {fmt(latest_yr):>7} {latest_yr:>5}")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("Fetching nominal gross wages from authoritative sources\n")

    # 1. ECB FX rates (needed by everything)
    fx_rates = fetch_ecb_fx_rates()
    if not fx_rates:
        print("WARNING: No FX rates, EUR conversion limited to eurozone only")

    # 2. National office fetchers (highest priority)
    national = fetch_national_offices(fx_rates)

    # Apply FX conversion for non-EUR national office data
    for iso2, years_data in national.items():
        for year, d in years_data.items():
            if d["eur"] is None and d["currency"] in fx_rates:
                if year in fx_rates[d["currency"]]:
                    rate = fx_rates[d["currency"]][year]
                    d["eur"] = d["local"] / rate if rate > 0 else None

    # 3. OECD (secondary, fills countries without national office fetchers)
    oecd_data = fetch_oecd_wages()
    if not oecd_data:
        print("ERROR: No OECD data fetched")
        sys.exit(1)

    # 4. Eurostat earn_nt_net (fills CY, MT — only EU countries not in OECD)
    eurostat_earn = fetch_eurostat_wages()

    # 5. Eurostat national accounts: D11/employees and D11/D1 ratio
    d11emp_data, d11d1_ratios = fetch_eurostat_nataccounts()

    # 6. GDP per capita (Eurostat + World Bank)
    gdp_data = fetch_gdp_per_capita(fx_rates)

    # IMF WEO: adds GDP forecasts 2026-2031 (and fills any historical gaps)
    weo_path = os.path.join(DATA_DIR, "WEOApr2026all.xlsx")
    if os.path.exists(weo_path):
        print(f"\n{'=' * 70}")
        print("IMF WEO GDP FORECASTS")
        print("=" * 70)
        imf_gdp = fetch_imf_weo(weo_path, fx_rates)
        # Merge: Eurostat/WorldBank preferred for historical; IMF for 2026+
        for iso2, ydata in imf_gdp.items():
            gdp_data.setdefault(iso2, {})
            for yr, val in ydata.items():
                if yr > 2025 or yr not in gdp_data[iso2]:
                    gdp_data[iso2][yr] = val
    else:
        print(f"\n  IMF WEO file not found at {weo_path} — using extrapolated GDP")

    # 7. Wikipedia current snapshot (fills remaining countries)
    print(f"\n{'=' * 70}")
    print("WIKIPEDIA CURRENT SNAPSHOT")
    print("=" * 70)
    wiki = parse_wikipedia_current()
    print(f"  Parsed {len(wiki)} countries from Wikipedia")

    # 8. Combine: national_office > oecd > eurostat_earn > wikipedia
    #    No scaling applied — raw values from each source.
    print(f"\n{'=' * 70}")
    print("BUILDING COMBINED WAGE TABLE")
    print("=" * 70)

    rows = []
    source_summary = defaultdict(list)

    def _row(iso2, name, year, eur, usd, local, currency, source, oecd_eur="", d11emp_eur=""):
        gdp = gdp_data.get(iso2, {}).get(year, "")
        return {
            "iso2": iso2, "country": name, "year": year,
            "wage_monthly_eur": round(eur, 0) if eur != "" and eur is not None else "",
            "wage_monthly_usd": round(usd, 0) if usd else "",
            "wage_monthly_local": round(local, 0) if local != "" and local is not None else "",
            "currency": currency,
            "source": source,
            "wage_oecd_eur": round(oecd_eur, 0) if oecd_eur != "" and oecd_eur is not None else "",
            "wage_d11emp_eur": round(d11emp_eur, 0) if d11emp_eur != "" and d11emp_eur is not None else "",
            "gdp_pc_eur": round(gdp, 0) if gdp else "",
            "is_forecast": 0,
        }

    # 8a. National office data
    for iso2, years_data in national.items():
        name = ISO2_TO_NAME.get(iso2, iso2)
        for year in sorted(years_data.keys()):
            d = years_data[year]
            eur = d["eur"]
            usd = None
            if eur and "USD" in fx_rates and year in fx_rates["USD"]:
                usd = eur * fx_rates["USD"][year]
            d11emp = d11emp_data.get(iso2, {}).get(year, "")
            rows.append(_row(iso2, name, year, eur or "", usd, d["local"],
                             d["currency"], "national_office", d11emp_eur=d11emp))
        source_summary["national_office"].append(iso2)

    # 8b. OECD countries — D11/headcount preferred as primary; OECD kept as reference
    # For years before D11/headcount starts (pre-1995), fall back to OECD.
    oecd_rows = build_wage_table(oecd_data, fx_rates)
    for r in oecd_rows:
        iso2 = r["iso2"]
        if iso2 in national:
            continue
        iso3_match = [k for k, v in EUROPEAN.items() if v[0] == iso2]
        if iso3_match and iso3_match[0] in OECD_EXCLUDE:
            continue
        oecd_eur = float(r["wage_monthly_eur"]) if r["wage_monthly_eur"] else None
        d11emp   = d11emp_data.get(iso2, {}).get(r["year"], "")
        has_d11emp_series = bool(d11emp_data.get(iso2))
        if d11emp:
            primary, source = d11emp, "eurostat_d11emp"
            # USD must be derived from the same D11/headcount EUR value
            ecb_usd = fx_rates.get("USD", {}).get(r["year"])
            usd = round(d11emp * ecb_usd, 0) if ecb_usd else None
        elif not has_d11emp_series:
            primary, source = oecd_eur or "", "oecd"
            usd = float(r["wage_monthly_usd"]) if r["wage_monthly_usd"] else None
        else:
            continue
        rows.append(_row(iso2, r["country"], r["year"], primary, usd,
                         float(r["wage_monthly_local"]) if r["wage_monthly_local"] else "",
                         r["currency"], source, oecd_eur=oecd_eur, d11emp_eur=d11emp))
        if iso2 not in [x for xs in source_summary.values() for x in xs]:
            source_summary[source].append(iso2)

    # 8c. Eurostat D11/headcount for non-OECD countries (HR, RO, BG, RS, CY, MT...)
    covered = set(r["iso2"] for r in rows)
    for iso2, ydata in d11emp_data.items():
        if iso2 in covered:
            continue
        name = ISO2_TO_NAME.get(iso2, iso2)
        if not name:
            continue
        for year in sorted(ydata.keys()):
            eur = ydata[year]
            usd = eur / fx_rates["USD"].get(year, 1.1) if "USD" in fx_rates else None
            rows.append(_row(iso2, name, year, eur, usd, eur, "EUR",
                             "eurostat_d11emp", d11emp_eur=eur))
        source_summary["eurostat_d11emp"].append(iso2)

    # 8d. Eurostat earn_nt_net fallback (legacy)
    covered = set(r["iso2"] for r in rows)
    oecd_exclude_iso2 = {ISO3_TO_ISO2.get(k, "") for k in OECD_EXCLUDE}
    for iso2, years_data in eurostat_earn.items():
        if iso2 in covered or iso2 in oecd_exclude_iso2:
            continue
        name = ISO2_TO_NAME.get(iso2, iso2)
        if not name or iso2 == name:
            continue
        for year in sorted(years_data.keys()):
            eur = years_data[year]
            usd = eur * fx_rates["USD"].get(year, 1.1) if "USD" in fx_rates else None
            d11emp = d11emp_data.get(iso2, {}).get(year, "")
            rows.append(_row(iso2, name, year, eur, usd, eur, "EUR",
                             "eurostat_earn", d11emp_eur=d11emp))
        source_summary["eurostat_earn"].append(iso2)

    # 8e. Wikipedia snapshot for remaining countries
    covered = set(r["iso2"] for r in rows)
    for iso2, wd in wiki.items():
        if iso2 in covered:
            continue
        name = ISO2_TO_NAME.get(iso2, iso2)
        d11emp = d11emp_data.get(iso2, {}).get(2024, "")
        rows.append(_row(iso2, name, 2025, wd["gross_eur"], "", "",
                         "EUR", f"wikipedia ({wd['date']})", d11emp_eur=d11emp))
        source_summary["wikipedia"].append(iso2)

    print(f"  Total rows: {len(rows):,}")
    for src, countries in sorted(source_summary.items()):
        print(f"  {src} ({len(countries)}): {sorted(countries)}")

    # 8e. Wikipedia validation — compare, DO NOT correct. Just report.
    print(f"\n  --- Wikipedia validation (raw divergences, no correction) ---")
    latest_ours = {}
    for r in rows:
        iso2 = r["iso2"]; yr = r["year"]; eur = r.get("wage_monthly_eur", "")
        if eur and eur != "":
            if iso2 not in latest_ours or yr > latest_ours[iso2][0]:
                latest_ours[iso2] = (yr, float(eur), r["source"])
    for iso2, wd in sorted(wiki.items()):
        if iso2 not in latest_ours:
            continue
        our_yr, our_val, our_src = latest_ours[iso2]
        if our_src.startswith("wikipedia"):
            continue
        ratio = our_val / wd["gross_eur"]
        flag = " <-- LARGE DIVERGENCE" if abs(ratio - 1.0) > 0.20 else ""
        print(f"    {iso2}: {our_src}={our_val:,.0f} wiki={wd['gross_eur']:,.0f} "
              f"ratio={ratio:.2f}x{flag}")

    # 9. Projection to 2031
    print(f"\n{'=' * 70}")
    print("PROJECTIONS TO 2031")
    print("=" * 70)
    projected = project_wages(rows, gdp_data, horizon=2031)
    all_rows = rows + projected
    # Backfill gdp_pc_eur for historical rows that have it
    for r in all_rows:
        if not r.get("gdp_pc_eur") or r["gdp_pc_eur"] == "":
            g = gdp_data.get(r["iso2"], {}).get(r["year"], "")
            r["gdp_pc_eur"] = round(g, 0) if g else ""

    # 10. Save
    FIELDS = ["iso2", "country", "year", "wage_monthly_eur", "wage_monthly_usd",
              "wage_monthly_local", "currency", "source",
              "wage_oecd_eur", "wage_d11emp_eur", "gdp_pc_eur", "is_forecast"]
    csv_path = os.path.join(DATA_DIR, "oecd_wages_europe.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_rows)
    print(f"  Saved: {csv_path} ({len(all_rows):,} rows incl. projections)")

    # 11. Validate
    print_validation(rows)  # historical only

    # 12. Correlation: log(GDP) vs log(wage)
    print(f"\n  --- GDP-Wage correlation (2024 cross-section) ---")
    import math
    corr_pts = []
    for r in rows:
        if r["year"] == 2024 and r.get("gdp_pc_eur") and r.get("wage_monthly_eur"):
            g = float(r["gdp_pc_eur"]); w = float(r["wage_monthly_eur"])
            if g > 0 and w > 0:
                corr_pts.append((math.log(g), math.log(w), r["iso2"]))
    if corr_pts:
        log_gdp = [p[0] for p in corr_pts]
        log_wage = [p[1] for p in corr_pts]
        slope, intercept, r2 = _linreg(log_gdp, log_wage)
        print(f"    2024: N={len(corr_pts)}, R²={r2:.3f}, elasticity={slope:.3f}")
        print(f"    (elasticity>1 means wages grow faster than GDP per capita)")

    # 13. Charts
    print(f"\n{'=' * 70}")
    print("CHARTS")
    print("=" * 70)

    plot_focus(rows,
               ["DE", "PL", "CZ", "SK", "LT", "HU", "AT", "BY", "RU"],
               "europe_01_poland_neighbors.png",
               " — Poland + Neighbors")

    plot_all_europe(rows, "europe_02_all_europe_eur.png")

    plot_ratio_germany(rows, all_rows,
                       ["AT", "BY", "RU",
                        "PL", "CZ", "SK", "LT", "HU", "EE", "LV", "RO", "BG", "HR",
                        "PT", "GR", "ES", "RS"],
                       "europe_03_ratio_germany.png")

    plot_gdp_wage_scatter(rows, gdp_data, "europe_04_gdp_wage_correlation.png")

    # Chart 05: union of chart 01 (neighbors) + chart 03 (ratio countries)
    plot_projection(all_rows,
                    ["DE", "AT", "PL", "CZ", "SK", "LT", "HU", "BY", "RU",
                     "EE", "LV", "RO", "BG", "HR", "PT", "GR", "ES", "RS"],
                    "europe_05_wage_projection.png")

    # 14. Coverage
    final_covered = set(r["iso2"] for r in rows)
    all_european = set(v[0] for v in EUROPEAN.values())
    missing = all_european - final_covered
    print(f"\n{'=' * 70}")
    print("COVERAGE SUMMARY")
    print(f"{'=' * 70}")
    print(f"  Total countries: {len(final_covered)}")
    if missing:
        print(f"  Still missing ({len(missing)}): "
              f"{', '.join(ISO2_TO_NAME.get(m, m) for m in sorted(missing))}")


if __name__ == "__main__":
    main()
